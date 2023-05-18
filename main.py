# Programa pra pesquisar preços de play5 e criar tabela com os valores destancando o mais barato

from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

# Configurar opções do webdriver
options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')
options.add_argument('--ignore-ssl-errors')
options.add_argument('--headless')

# Configurar o serviço do webdriver
service = Service(ChromeDriverManager().install())

# Instanciar o webdriver
driver = webdriver.Chrome(service=service, options=options)

# Criar workbook e sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Adicionar cabeçalhos da tabela
sheet['A1'] = 'Produto'
sheet['B1'] = 'Amazon'
sheet['C1'] = 'Kabum'

urls = {
    "Play5": [
        "https://www.amazon.com.br/PlayStation-CFI-1214B01X-Console-PlayStation%C2%AE5-Digital/dp/B0BNSR6S9Z/ref=asc_df_B0BNSR6S9Z/?tag=googleshopp00-20&linkCode=df0&hvadid=426445053297&hvpos=&hvnetw=g&hvrand=17901529703630833800&hvpone=&hvptwo=&hvqmt=&hvdev=c&hvdvcmdl=&hvlocint=&hvlocphy=1031630&hvtargid=pla-1943919044802&psc=1",
        "https://www.kabum.com.br/produto/238670/console-sony-playstation-5-edicao-digital-1214b"
    ],
}

for i, (produto, urls_produto) in enumerate(urls.items(), start=2):
    sheet.cell(row=i, column=1, value=produto)
    precos = []
    for j, url in enumerate(urls_produto, start=2):
        driver.get(url)
        time.sleep(5)
        try:
            if "www.amazon.com.br" in url:
                preco_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, '.a-offscreen'))
                )
                preco_texto = preco_element.get_attribute('innerText')
                preco_numerico = preco_texto.replace('R$', '').replace('.', '').replace(',', '').strip()
                try:
                    preco = float(preco_numerico) * 100  # Converter para centavos de reais
                    print(f"Preço do produto {produto} em {url}: R$ {preco / 100:.2f}")
                    precos.append(preco)
                except ValueError:
                    print(f"Erro ao converter preço para float: {preco_numerico}")
                    continue  # Pula para a próxima iteração do loop

            elif "www.kabum.com.br" in url:
                preco_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="blocoValores"]/div[2]/div[1]/h4'))
                )
                preco_texto = preco_element.get_attribute('innerText')
                preco_numerico = preco_texto.replace('R$', '').replace('.', '').replace(',', '').strip()
                try:
                    preco = float(preco_numerico) * 100  # Converter para centavos de reais
                    print(f"Preço do produto {produto} em {url}: R$ {preco / 100:.2f}")
                    precos.append(preco)
                except ValueError:
                    print(f"Erro ao converter preço para float: {preco_numerico}")
                    continue  # Pula para a próxima iteração do loop

        except Exception as e:
            print(f"Erro ao buscar preço no site {url}: {e}")
            continue  # Pula para a próxima iteração do loop

    if precos:
        preco_min = min(precos)
        for j, preco in enumerate(precos, start=2):
            sheet.cell(row=i, column=j, value=preco)
            if preco == preco_min:
                sheet.cell(row=i, column=j).style = 'Good'

# Salvar planilha
workbook.save('precos.xlsx')

driver.quit()

print('Preços salvos com sucesso!')
